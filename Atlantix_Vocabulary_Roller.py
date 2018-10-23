#coding: utf-8
from openpyxl import load_workbook
import sys, os, time
import termios, threading
try:
    from AppKit import NSSpeechSynthesizer
    speaker = NSSpeechSynthesizer.alloc().init()
    USE_TTS = True
except:
    print("!> Cannot init TTS")
    USE_TTS = False
try: input = raw_input
except NameError: pass

def press_any_key_exit():
  # 获取标准输入的描述符
  fd = sys.stdin.fileno()
 
  # 获取标准输入(终端)的设置
  old_ttyinfo = termios.tcgetattr(fd)
 
  # 配置终端
  new_ttyinfo = old_ttyinfo[:]
 
  # 使用非规范模式(索引3是c_lflag 也就是本地模式)
  new_ttyinfo[3] &= ~termios.ICANON
  # 关闭回显(输入不会被显示)
  new_ttyinfo[3] &= ~termios.ECHO
 
  # 输出信息
  # sys.stdout.write(msg)
  # sys.stdout.flush()

  # 使设置生效
  termios.tcsetattr(fd, termios.TCSANOW, new_ttyinfo)
  # 从终端读取
  c = os.read(fd, 7)
 
  # 还原终端设置
  termios.tcsetattr(fd, termios.TCSANOW, old_ttyinfo)

  return c

def wait_or_until(obj, name, t):
    cnt = 0
    num = int(t / 0.1)
    while cnt < num:
        if getattr(obj, name): break
        time.sleep(0.1)
        cnt += 1

class CMDControl(object):
    def __init__(self, database, last_pos):
        self.last_pos = last_pos
        self.database = database
        self.pronouce_list = []
        self.show = False
        self.stop = False
        self.pause = False; self.resume = True
        self.thr = threading.Thread(target=CMDControl.worker, args=(self.database, self))
        if USE_TTS:
            self.speak_thr = threading.Thread(target=CMDControl.pronouce, args=(speaker, self))
    
    def start_thread(self):
        self.thr.setDaemon(True)
        self.thr.start()
        if USE_TTS:
            self.speak_thr.setDaemon(True)
            self.speak_thr.start()

    @staticmethod
    def pronouce(speaker, obj):
        while not obj.stop:
            if len(obj.pronouce_list) > 0:
                s = obj.pronouce_list.pop()
                speaker.startSpeakingString_(s)
            time.sleep(0.5)

    @staticmethod
    def worker(database, obj):
        while obj.last_pos < len(database):
            print("")
            fraction = int((float(obj.last_pos) / len(database)) * 40)
            fraction_str = "[%d/%d]" % (obj.last_pos, len(database))
            print("=" * fraction + "." * (40 - fraction) + fraction_str)
            i = obj.last_pos

            print("=> " + database[i][0])
            if USE_TTS:
                obj.pronouce_list.append(database[i][0])
            wait_or_until(obj, "show", 2.0); obj.show = False

            if obj.pause: wait_or_until(obj, "resume", 100.0); obj.show = False

            print("- " + database[i][1])
            wait_or_until(obj, "show", 1.0); obj.show = False

            if obj.pause: wait_or_until(obj, "resume", 100.0); obj.show = False

            os.system("clear")
            obj.last_pos += 1
            if obj.stop: break

        return True

# excel = load_workbook("./再要你命3000电子版（大字版）.xlsx")
# table = excel.get_sheet_by_name(u'3000\u5355\u8bcd\u8868')   #通过表名获取  
# word_col = 2; meaning_col = 3
excel = load_workbook("再要你命3000自动计算版.xlsx")
table = excel.get_sheet_by_name('Sheet1')
word_col = 1; meaning_col = 2
n_rows = table.max_row   #获取行数
n_cols = table.max_column    #获取列数

database = []
hardlist = []

for row in range(1, n_rows + 1):
    word = table.cell(row=row,column=word_col).value
    meaning = table.cell(row=row,column=meaning_col).value
    database.append((word, meaning.replace(u'；', '\n- ').replace(u'; ', '\n- ').replace(u';', '\n- ').encode("utf-8")))

try:
    save = open("save.txt", "r")
    line = save.readline().strip().split(" ")
    last_pos, hard_last_pos = [int(item) for item in line]
    try:
        hardlist = [int(item) for item in save.readline().strip().split(" ")]
    except:
        print("!> No hard list")
except Exception as e:
    print(e)
    last_pos = 0
    hard_last_pos = 0

hard_database = [(database[i][0], database[i][1]) for i in hardlist]

print("| Automatic vocabulary roller, by Jianjin Xu")
print("| 2018-10-14")
print("| Usage:")
print("=> [e]xit [p]ause")
print("=> mark [h]ard, or [d]elete")
print("=> [n]ext, or [b] for previous")

print("Roll 3000 or roll hard list(1/2)? ")
rtype = press_any_key_exit()
if rtype == '1':
    db = database
    pos = last_pos
    print("=> Rolling 3000 from %d" % last_pos)
elif rtype == '2':
    db = hard_database
    pos = hard_last_pos
    if pos >= len(hard_database): pos = 0
    print("=> Rolling hard list from %d" % hard_last_pos)

cmd_ctrl = CMDControl(db, pos)
cmd_ctrl.start_thread()

while True:
    try:
        c = press_any_key_exit()
    except KeyboardInterrupt:
        print("!> Press e to exit")
    
    # print(c)
    if c == 'e':
        cmd_ctrl.stop = True
        break
    elif c == 'h':
        print("=> Mark %s as hard" % database[cmd_ctrl.last_pos][0])
        hardlist.append(cmd_ctrl.last_pos)
    elif c == 'n':
        cmd_ctrl.show = True
    elif c == 'b':
        cmd_ctrl.last_pos -= 2
    elif c == 'd':
        print("=> Delete %s from hard list" % hard_database[cmd_ctrl.last_pos][0])
        hardlist[cmd_ctrl.last_pos] = -1
    elif c == 'p':
        if cmd_ctrl.pause:
            print("=> Resume")
            cmd_ctrl.pause = False
            cmd_ctrl.resume = True
        else:
            print("=> Pause")
            cmd_ctrl.pause = True
            cmd_ctrl.resume = False
pos = cmd_ctrl.last_pos
if rtype == '1': last_pos = pos
elif rtype == '2': hard_last_pos = pos

save = open("save.txt", "w")
save.write("%d %d\n" % (last_pos, hard_last_pos))
s = ""
for i in hardlist:
    if i > 0: s += "%d " % i
save.write(s)
save.close()